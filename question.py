import tkinter as tk
from tkinter import filedialog, messagebox
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def get_next_question_number():
    """
    Connect to question_bank.db, ensure the table exists, and return the next available question number.
    """
    try:
        conn = sqlite3.connect("question_bank.db")
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS "computer and communication systems" (
                "Question No." INTEGER PRIMARY KEY,
                "Question" TEXT,
                "Option A" TEXT,
                "Option B" TEXT,
                "Option C" TEXT,
                "Option D" TEXT,
                "Answer" TEXT
            )
        ''')
        cur.execute('SELECT MAX("Question No.") FROM "computer and communication systems"')
        result = cur.fetchone()[0]
        conn.close()
        if result is None:
            return 1
        else:
            return result + 1
    except Exception as e:
        # In case of any error, default to 1
        return 1

def download_template():
    """
    Create a preformatted Excel file with headers, prefilled first data row using the next question number from the DB,
    and data validation for the Answer column.
    """
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define the headers as specified
    headers = ["Question No.", "Question", "Option A", "Option B", "Option C", "Option D", "Answer"]
    ws.append(headers)

    # Get the next available question number from the database
    next_question_no = get_next_question_number()

    # Pre-fill the first row of data with the next question number and default answer "A1"
    ws.append([next_question_no, "", "", "", "", "", "A1"])

    # Add data validation to the Answer column (Column G) for rows 2 to 100
    dv = DataValidation(type="list", formula1='"A1,A2,A3,A4"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("G2:G100")

    # Open a save dialog to let the user choose the file location and name
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Template As"
    )

    if file_path:
        try:
            wb.save(file_path)
            messagebox.showinfo("Success", "Template downloaded successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Error saving file: {e}")

def upload_excel():
    """
    Allow the user to select an Excel file, validate the contents, and insert/update the questions into the
    'computer and communication systems' table in question_bank.db.
    """
    # Open a file dialog to select the Excel file
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")],
        title="Select Quiz Excel File"
    )
    if not file_path:
        return

    try:
        # Load the workbook and select the active worksheet
        wb = load_workbook(file_path)
        ws = wb.active

        # Allowed answer values
        valid_answers = {"A1", "A2", "A3", "A4"}

        # Connect to the SQLite database (or create it if it doesn't exist)
        conn = sqlite3.connect("question_bank.db")
        cur = conn.cursor()

        # Create the table if it doesn't exist already.
        cur.execute('''
            CREATE TABLE IF NOT EXISTS "computer and communication systems" (
                "Question No." INTEGER PRIMARY KEY,
                "Question" TEXT,
                "Option A" TEXT,
                "Option B" TEXT,
                "Option C" TEXT,
                "Option D" TEXT,
                "Answer" TEXT
            )
        ''')

        rows_added = 0
        # Iterate over the rows in the worksheet, skipping the header row (assumed to be row 1)
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue

            q_no, question, opt_a, opt_b, opt_c, opt_d, answer = row

            # Validate Question No.
            if q_no is None:
                messagebox.showerror("Error", f"Missing Question No. in row {idx}.")
                return
            try:
                q_no = int(q_no)
            except Exception:
                messagebox.showerror("Error", f"Invalid Question No. in row {idx}. It must be an integer.")
                return

            # Validate Answer value
            if answer not in valid_answers:
                messagebox.showerror("Error", f"Invalid answer '{answer}' in row {idx}. Allowed values are: A1, A2, A3, A4.")
                return

            # Insert or replace the record in the table
            cur.execute('''
                INSERT OR REPLACE INTO "computer and communication systems"
                ("Question No.", "Question", "Option A", "Option B", "Option C", "Option D", "Answer")
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (q_no, question, opt_a, opt_b, opt_c, opt_d, answer))
            rows_added += 1

        conn.commit()
        conn.close()
        messagebox.showinfo("Success", f"Upload successful! {rows_added} questions added/updated.")
    except Exception as e:
        messagebox.showerror("Error", f"Error uploading file: {e}")

# Set up the main Tkinter window
root = tk.Tk()
root.title("Quiz Creator Tool")
root.geometry("400x220")
root.resizable(False, False)

# Title label
title_label = tk.Label(root, text="Quiz Creator Tool", font=("Helvetica", 16, "bold"))
title_label.pack(pady=10)

# Download template button
download_btn = tk.Button(root, text="Download Template", command=download_template, width=25, height=2)
download_btn.pack(pady=10)

# Upload Excel file button
upload_btn = tk.Button(root, text="Upload Quiz Excel", command=upload_excel, width=25, height=2)
upload_btn.pack(pady=10)

root.mainloop()
