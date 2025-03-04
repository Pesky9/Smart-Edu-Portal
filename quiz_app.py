from tkinter import *
from tkinter import messagebox, ttk,filedialog
import csv
import sqlite3
import os
import random
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageTk  
cor_wron = []
quiz_score = {}
score = 0

def database_conn():
    global conn, cursor
    conn = sqlite3.connect("user_data.db")
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS `user` (user_id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, firstname TEXT, surname TEXT, username TEXT, password TEXT)")

def login():
    database_conn()
    if username.get() == "" or password.get() == "":
        messagebox.showerror("Empty Fields", "Please enter both username and password!")
    else:
        cursor.execute("SELECT * FROM `user` WHERE `username` = ? AND `password` = ?", (username.get(), password.get()))
        if cursor.fetchone() is not None:
            home()
        else:
            messagebox.showerror("Invalid User", "The Username or Password is incorrect. Try Again")
            username.delete(0, END)
            password.delete(0, END)
    cursor.close()
    conn.close()

def admin_log():
    def log_back():
        up.destroy()
        log.deiconify()
        username.delete(0, END)
        password.delete(0, END)

    admin_username = 'Admin'
    admin_pass = '1234'

    def signup_db():
        up.withdraw()
        global ques_e, quest, op_a_e, op_b_e, op_c_e, op_d_e, ans_e
        ques_e = StringVar()
        quest = StringVar()
        op_a_e = StringVar()
        op_b_e = StringVar()
        op_c_e = StringVar()
        op_d_e = StringVar()
        ans_e = StringVar()
        database_conn()
        if username1.get() == "" or password1.get() == "":
            messagebox.showerror("Empty Fields", "Please enter both username and password!")
        else:
            if username1.get() == admin_username and password1.get() == admin_pass:
                messagebox.showinfo("Logged In", "Successfully Logged In")
                def quiz_data(name):
                    adminportal = Tk()
                    adminportal.geometry('1110x600')
                    adminportal.configure(bg="#f8f8f8")
                    treev2 = ttk.Treeview(adminportal, selectmode='browse')
                    treev2.place(height=370, width=1110, x=0, y=30)
                    vsb = ttk.Scrollbar(adminportal, orient="vertical", command=treev2.yview)
                    vsb.place(x=1093, y=31, height=368)
                    treev2.configure(yscrollcommand=vsb.set)
                    hsb = ttk.Scrollbar(adminportal, orient="horizontal", command=treev2.xview)
                    hsb.place(x=1, y=379, width=1092, height=20)
                    treev2.configure(xscrollcommand=hsb.set)
                    treev2["columns"] = ("1", "2", "3", "4", "5", "6", "7")
                    treev2['show'] = 'headings'
                    treev2.column("1", width=100, anchor='c')
                    treev2.column("2", width=200, anchor='c')
                    treev2.column("3", width=100, anchor='c')
                    treev2.column("4", width=100, anchor='c')
                    treev2.column("5", width=100, anchor='c')
                    treev2.column("6", width=100, anchor='c')
                    treev2.column("7", width=120, anchor='c')
                    treev2.heading("1", text="Question No")
                    treev2.heading("2", text="Question")
                    treev2.heading("3", text="Option A")
                    treev2.heading("4", text="Option B")
                    treev2.heading("5", text="Option C")
                    treev2.heading("6", text="Option D")
                    treev2.heading("7", text="Answer")
                    mydb2 = sqlite3.connect("question_bank.db")
                    mycursor2 = mydb2.cursor()
                    treev2.delete(*treev2.get_children())
                    mycursor2.execute(f"SELECT * FROM [{name}] LIMIT 100;")
                    data_2 = mycursor2.fetchall()
                    indexer = 1
                    for value in data_2:
                        treev2.insert("", 'end', text="L" + str(indexer),
                                      values=(value[0], value[1], value[2], value[3], value[4], value[5], value[6]))
                        indexer += 1

                    def update_ques():
                        mycursor2.execute(
                            f"UPDATE [{name}] SET [Question] =?, [Option A] =?, [Option B] =?, [Option C] =?, [Option D] =?, [Answer] =? WHERE [Question No.] =?",
                            (quest.get(), op_a_e.get(), op_b_e.get(), op_c_e.get(), op_d_e.get(), ans_e.get(), q_no_e.get()))
                        mydb2.commit()
                        adminportal.withdraw()
                        adminportal.after(1, lambda: quiz_data(name))

                    def delete_ques():
                        mycursor2.execute(f"DELETE FROM [{name}] WHERE [Question No.] = {q_no_e.get()}")
                        mydb2.commit()
                        adminportal.withdraw()
                        adminportal.after(1, lambda: quiz_data(name))

                    def add_question():
                        new_win = Tk()
                        new_win.resizable(0, 0)
                        new_win.title('Add Question')
                        def submit_quest():
                            if (question_e.get() == '' or 
                                option_a_e.get() == '' or 
                                option_b_e.get() == '' or 
                                option_c_e.get() == '' or 
                                option_d_e.get() == '' or 
                                answer_e.get() == ''):
                                messagebox.showerror('Error', 'Please enter all the values')
                            else:
                                mycursor2.execute(
                                    f"INSERT INTO [{name}] ([Question],[Option A],[Option B],[Option C],[Option D],[Answer]) VALUES(?,?,?,?,?,?)",
                                    (question_e.get(), option_a_e.get(), option_b_e.get(), option_c_e.get(), option_d_e.get(), answer_e.get()))
                                mydb2.commit()
                                new_win.withdraw()
                                adminportal.withdraw()
                                adminportal.after(1, lambda: quiz_data(name))
                        question_head = Label(new_win, text='Add a Question', font=("ARIAL", 16))
                        question_head.grid(row=0, column=0, padx=10, pady=10)
                        question = Label(new_win, text='Question', font=("ARIAL", 16))
                        question.grid(row=1, column=0, padx=10, pady=10)
                        question_e = Entry(new_win, font=("ARIAL", 16))
                        question_e.grid(row=1, column=1)
                        option_a = Label(new_win, text='Option A', font=("ARIAL", 16))
                        option_a.grid(row=2, column=0, padx=10, pady=10)
                        option_a_e = Entry(new_win, font=("ARIAL", 16))
                        option_a_e.grid(row=2, column=1)
                        option_b = Label(new_win, text='Option B', font=("ARIAL", 16))
                        option_b.grid(row=3, column=0, padx=10, pady=10)
                        option_b_e = Entry(new_win, font=("ARIAL", 16))
                        option_b_e.grid(row=3, column=1)
                        option_c = Label(new_win, text='Option C', font=("ARIAL", 16))
                        option_c.grid(row=4, column=0, padx=10, pady=10)
                        option_c_e = Entry(new_win, font=("ARIAL", 16))
                        option_c_e.grid(row=4, column=1)
                        option_d = Label(new_win, text='Option D', font=("ARIAL", 16))
                        option_d.grid(row=5, column=0, padx=10, pady=10)
                        option_d_e = Entry(new_win, font=("ARIAL", 16))
                        option_d_e.grid(row=5, column=1)
                        answer = Label(new_win, text='Answer', font=("ARIAL", 16))
                        answer.grid(row=6, column=0, padx=10, pady=10)
                        answer_e = Entry(new_win, font=("ARIAL", 16))
                        answer_e.grid(row=6, column=1, padx=20)
                        submit = Button(new_win, text='Add Question', command=submit_quest)
                        submit.config(font=("ARIAL", 12))
                        submit.grid(row=7, column=1, pady=15)

                    def bulk_upload():
                        option_win = Toplevel(adminportal)
                        option_win.title("Bulk Upload Options")
                        option_win.geometry("300x150")
                        def get_next_question_no():
                            try:                      

                                mycursor2.execute(f"SELECT MAX([Question No.]) FROM [{name}]")
                                result = mycursor2.fetchone()[0]
                                return 1 if result is None else result + 1
                            except Exception as e:
                                return 1

                        def download_template_option():
                            from openpyxl import Workbook
                            from openpyxl.worksheet.datavalidation import DataValidation
                            wb = Workbook()
                            ws = wb.active
                            headers = ["Question No.", "Question", "Option A", "Option B", "Option C", "Option D", "Answer"]
                            ws.append(headers)
                            next_q_no = get_next_question_no()
                            ws.append([next_q_no, "", "", "", "", "", "A1"])
                            dv = DataValidation(type="list", formula1='"A1,A2,A3,A4"', allow_blank=False)
                            ws.add_data_validation(dv)
                            dv.add("G2:G100")
                            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                                     filetypes=[("Excel files", "*.xlsx")],
                                                                     title="Save Template As")
                            if file_path:
                                try:
                                    wb.save(file_path)
                                    messagebox.showinfo("Success", "Template downloaded successfully!")
                                except Exception as e:
                                    messagebox.showerror("Error", f"Error saving file: {e}")
                            option_win.destroy()

                        def upload_template_option():
                            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")],
                                                                   title="Select Bulk Upload Excel File")
                            if not file_path:
                                return
                            try:
                                from openpyxl import load_workbook
                                wb = load_workbook(file_path)
                                ws = wb.active
                                valid_answers = {"A1", "A2", "A3", "A4"}
                                rows_added = 0
                                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                                    if not any(row):
                                        continue
                                    q_no, question, op_a, op_b, op_c, op_d, answer = row
                                    if q_no is None:
                                        messagebox.showerror("Error", f"Missing Question No. in row {idx}.")
                                        return
                                    try:
                                        q_no = int(q_no)
                                    except Exception:
                                        messagebox.showerror("Error", f"Invalid Question No. in row {idx}. It must be an integer.")
                                        return
                                    if answer not in valid_answers:
                                        messagebox.showerror("Error", f"Invalid answer '{answer}' in row {idx}. Allowed values are: A1, A2, A3, A4.")
                                        return
                                    mycursor2.execute(
                                        f"INSERT OR REPLACE INTO [{name}] ([Question No.], [Question], [Option A], [Option B], [Option C], [Option D], [Answer]) VALUES (?,?,?,?,?,?,?)",
                                        (q_no, question, op_a, op_b, op_c, op_d, answer))
                                    rows_added += 1
                                mydb2.commit()
                                messagebox.showinfo("Bulk Upload Success", f"Bulk upload successful! {rows_added} questions added/updated.")
                                option_win.destroy()
                                adminportal.withdraw()
                                adminportal.after(1, lambda: quiz_data(name))
                            except Exception as e:
                                messagebox.showerror("Error", f"Error during bulk upload: {e}")
                        
                        Label(option_win, text="Choose Bulk Upload Option", font=("Arial", 12)).pack(pady=10)
                        Button(option_win, text="Download Template", font=("Arial", 12), command=download_template_option).pack(pady=5)
                        Button(option_win, text="Upload Template", font=("Arial", 12), command=upload_template_option).pack(pady=5)
               
                    ad_q = Button(adminportal, text='Add Question', command=add_question)
                    ad_q.config(bg='white', font=("ARIAL", 12))
                    ad_q.place(x=30, y=420)
                    ed_q = Button(adminportal, text='Update Question', command=update_ques)
                    ed_q.config(bg='white', font=("ARIAL", 12))
                    ed_q.place(x=280, y=420)
                    de_q = Button(adminportal, text='Delete Question', command=delete_ques)
                    de_q.config(bg='white', font=("ARIAL", 12))
                    de_q.place(x=510, y=420)
                    bulk_btn = Button(adminportal, text='Bulk Upload', command=bulk_upload)
                    bulk_btn.config(bg='white', font=("ARIAL", 12))
                    bulk_btn.place(x=750, y=420)
                    
                    q_no = Label(adminportal, text='Question No.', state='disabled')
                    q_no.config(bg='white', font=("ARIAL", 12))
                    q_no.place(x=30, y=480)
                    q_no_e = Entry(adminportal, width=13)
                    q_no_e.config(bg='white', font=("ARIAL", 12))
                    q_no_e.place(x=130, y=480)
                    ques = Label(adminportal, text='Question')
                    ques.config(bg='white', font=("ARIAL", 12))
                    ques.place(x=280, y=480)
                    quest = Entry(adminportal, width=13)
                    quest.config(bg='white', font=("ARIAL", 12))
                    quest.place(x=360, y=480)
                    op_a = Label(adminportal, text='Option A')
                    op_a.config(bg='white', font=("ARIAL", 12))
                    op_a.place(x=490, y=480)
                    op_a_e = Entry(adminportal, width=13)
                    op_a_e.config(bg='white', font=("ARIAL", 12))
                    op_a_e.place(x=570, y=480)
                    op_b = Label(adminportal, text='Option B')
                    op_b.config(bg='white', font=("ARIAL", 12))
                    op_b.place(x=700, y=480)
                    op_b_e = Entry(adminportal, width=13)
                    op_b_e.config(bg='white', font=("ARIAL", 12))
                    op_b_e.place(x=780, y=480)
                    op_c = Label(adminportal, text='Option C')
                    op_c.config(bg='white', font=("ARIAL", 12))
                    op_c.place(x=280, y=530)
                    op_c_e = Entry(adminportal, width=13)
                    op_c_e.config(bg='white', font=("ARIAL", 12))
                    op_c_e.place(x=360, y=530)
                    op_d = Label(adminportal, text='Option D')
                    op_d.config(bg='white', font=("ARIAL", 12))
                    op_d.place(x=490, y=530)
                    op_d_e = Entry(adminportal, width=13)
                    op_d_e.config(bg='white', font=("ARIAL", 12))
                    op_d_e.place(x=570, y=530)
                    ans = Label(adminportal, text='Answer')
                    ans.config(bg='white', font=("ARIAL", 12))
                    ans.place(x=707, y=530)
                    ans_e = Entry(adminportal, width=13)
                    ans_e.config(bg='white', font=("ARIAL", 12))
                    ans_e.place(x=780, y=530)
                    
                    def getrow(event):
                        q_no_e['state'] = 'normal'
                        q_no_e.delete(0, END)
                        q_no_e['state'] = 'disabled'
                        quest.delete(0, END)
                        op_a_e.delete(0, END)
                        op_b_e.delete(0, END)
                        op_c_e.delete(0, END)
                        op_d_e.delete(0, END)
                        ans_e.delete(0, END)
                        q_no_e['state'] = 'normal'
                        item = treev2.item(treev2.focus())
                        q_no_e.insert(END, item['values'][0])
                        q_no_e['state'] = 'disabled'
                        quest.insert(END, item['values'][1])
                        op_a_e.insert(END, item['values'][2])
                        op_b_e.insert(END, item['values'][3])
                        op_c_e.insert(END, item['values'][4])
                        op_d_e.insert(END, item['values'][5])
                        ans_e.insert(END, item['values'][6])
                    treev2.bind('<Double-Button-1>', getrow)
                    adminportal.mainloop()
                
                def view():
                    root2 = Toplevel()
                    root2.title('Quiz Subject')
                    root2.resizable(0, 0)
                    lbl = Label(root2, text='Pick a Subject', font=("HELVETICA", 15))
                    lbl.grid(row=0, column=0, columnspan=2, padx=40, pady=20)
                    con2 = sqlite3.connect("question_bank.db")
                    cursor2 = con2.cursor()
                    cursor2.execute("SELECT name FROM sqlite_master WHERE type='table';")
                    con2.commit()
                    aa = cursor2.fetchall()
                    cc = []
                    for i in aa:
                        cc.append(list(i))
                    data2 = cc
                    for i in range(len(data2)):
                        data2[i][0] = str(data2[i][0])
                    x = 0
                    for i in range(len(data2)):
                        ll = Label(root2, text=f'{data2[i][0].title()}', font=("HELVETICA", 15))
                        ll.grid(row=x, column=0, columnspan=2, padx=40, pady=20)
                        x += 1
                    y = 0
                    for j in range(len(data2)):
                        lla = Button(root2, text='Edit', width=5, font=("HELVETICA", 14),
                                    command=lambda name=f'{data2[j][0]}': quiz_data(name))
                        lla.grid(row=y, column=3)
                        y += 1
                    def del_tab(name):
                        if messagebox.askyesno("Info", "Are you sure you want to delete this subject?"):
                            query = f"DROP TABLE [{name}];"
                            root2.withdraw()
                            con2.execute(query)
                            root2.after(1000, view())
                            root2.deiconify()
                    z = 0
                    for k in range(len(data2)):
                        llal = Button(root2, text='Delete', width=5, font=("HELVETICA", 14),
                                    command=lambda name=f'{data2[k][0].title()}': del_tab(name))
                        llal.grid(row=z, column=4, padx=50)
                        z += 1
                    def add_sub():
                        root2.withdraw()
                        def proceed():
                            query = f"""CREATE TABLE "{quiz.get()}" (
                                                        "Question No."	INTEGER,
                                                        "Question"	TEXT,
                                                        "Option A"	TEXT,
                                                        "Option B"	TEXT,
                                                        "Option C"	TEXT,
                                                        "Option D"	TEXT,
                                                        "Answer"	TEXT,
                                                        PRIMARY KEY("Question No.")
                            );"""
                            cursor2.execute(query)
                            add_quiz.withdraw()
                            root2.after(1000, view())
                            root2.deiconify()
                        add_quiz = Tk()
                        add_quiz.geometry('410x300')
                        quiz_label = Label(add_quiz, text="Add a Subject", font=("HELVETICA", 15))
                        quiz_label.place(x=145, y=125)
                        quiz = Entry(add_quiz, width=30)
                        quiz.config(font=("ARIAL", 12))
                        quiz.place(x=70, y=175)
                        quiz_btn = Button(add_quiz, text='Proceed', width=10, font=("HELVETICA", 14),
                                        command=lambda: proceed())
                        quiz_btn.place(x=145, y=220)
                        add_quiz.mainloop()
                    llala = Button(root2, text='Add a Subject', width=20, font=("HELVETICA", 14),
                                command=lambda: add_sub())
                    llala.grid(row=z + 2, column=1, pady=20)
                    root2.mainloop()
                
                view()
            else:
                messagebox.showerror("Error", "Incorrect username or password!")
                admin_log()
    
    up = Tk()
    up.title('Sign Up')
    up.geometry("600x450")  # Set width and height of the window
    up.resizable(0, 0)  # Disable window resizing
    up.configure(bg="#e6f2ff")  # Light blue background color

# Create a frame for better layout
    frame = Frame(up, bg="#ffffff", padx=20, pady=20, relief=RAISED, bd=5)
    frame.pack(pady=20)

# Header label
    head = Label(frame, text='PROJECT QUIZ', font=('Helvetica', 21, 'bold'), bg="#ffffff", fg="#003366")
    head.grid(row=0, column=0, columnspan=2, pady=10)

# Username label and entry
    u_lbl = Label(frame, text='Username:', font=("Arial", 12, 'bold'), bg="#ffffff", fg="#333333")
    u_lbl.grid(row=1, column=0, sticky="w", pady=5)
    username1 = Entry(frame, width=25, font=("Arial", 12))
    username1.grid(row=1, column=1, pady=5)

# Password label and entry
    pa_lbl = Label(frame, text='Password:', font=("Arial", 12, 'bold'), bg="#ffffff", fg="#333333")
    pa_lbl.grid(row=2, column=0, sticky="w", pady=5)
    password1 = Entry(frame, width=25, show='*', font=("Arial", 12))
    password1.grid(row=2, column=1, pady=5)

# Sign-up button
    signup_btn = Button(frame, text='Sign Up', font=('Helvetica', 14, 'bold'), width=15, bg="#008CBA", fg="white",
                    bd=3, relief=RAISED, command=lambda: print("Sign up clicked"))
    signup_btn.grid(row=3, column=0, columnspan=2, pady=10)

# "Not an admin?" label
    dont = Label(frame, text="Not an admin?", font=("Helvetica", 12, 'italic'), bg="#ffffff", fg="#555555")
    dont.grid(row=4, column=0, columnspan=2, pady=5)

# Login as user button
    login_btn = Button(frame, text='Login as User', font=("Helvetica", 12, 'bold'), width=12, bg="#4CAF50", fg="white",
                   bd=3, relief=RAISED, command=lambda: print("Login as user clicked"))
    login_btn.grid(row=5, column=0, columnspan=2, pady=5)

# Run the application
    up.mainloop()


def signup():
    def log_back():
        up.destroy()
        log.deiconify()
        username.delete(0, END)
        password.delete(0, END)
    def signup_db():
        database_conn()
        if fname.get() == "" or sname.get() == "" or username1.get() == "" or password1.get() == "":
            messagebox.showerror("Empty Fields", "Please enter both userneame and password!")
        else:
            cursor.execute("SELECT * FROM `user` WHERE `username` = ?", (username1.get(),))
            if cursor.fetchone() is not None:
                messagebox.showerror("User Exists", "A User with this Username Already exists.\nTry new username!")
                username.set("")
            else:
                cursor.execute("INSERT INTO `user` (firstname, surname, username, password) VALUES(?, ?, ?, ?)",
                               (fname.get(), sname.get(), username1.get(), password1.get()))
                conn.commit()
        cursor.close()
        conn.close()
        messagebox.showinfo('Successful!', 'You are successfully registered!')
        up.destroy()
        log.deiconify()
    log.withdraw()
    up = Toplevel()
    up.title('Sign Up')
    up.resizable(0, 0)
    up.configure(bg="#f8f8f8")
    head = Label(up, text='PROJECT QUIZ', font=('HELVETICA', 21, 'bold'), bg="#f8f8f8")
    head.grid(row=0, column=0, padx=10, pady=10, columnspan=2)
    f_lbl = Label(up, text='First Name', font=("HELVETICA", 12), bg="#f8f8f8")
    f_lbl.grid(row=1, column=0, padx=10, pady=10)
    fname = Entry(up, width=20, font=("ARIAL", 12))
    fname.grid(row=1, column=1, pady=10, padx=(0, 40))
    s_lbl = Label(up, text='Surname', font=("HELVETICA", 12), bg="#f8f8f8")
    s_lbl.grid(row=2, column=0, padx=10, pady=10)
    sname = Entry(up, width=20, font=("ARIAL", 12))
    sname.grid(row=2, column=1, pady=10, padx=(0, 40))
    u_lbl = Label(up, text='Username', font=("ARIAL", 12), bg="#f8f8f8")
    u_lbl.grid(row=3, column=0, padx=10, pady=10)
    username1 = Entry(up, width=20, font=("ARIAL", 12))
    username1.grid(row=3, column=1, pady=10, padx=(0, 40))
    pa_lbl = Label(up, text='Password', font=("ARIAL", 12), bg="#f8f8f8")
    pa_lbl.grid(row=4, column=0, padx=10, pady=10)
    password1 = Entry(up, width=20, show='*', font=("ARIAL", 12))
    password1.grid(row=4, column=1, pady=10, padx=(0, 40))
    signup_btn = Button(up, text='Sign Up', font=('HELVETICA', 16), width=12, bd=4, relief=RAISED, command=signup_db)
    signup_btn.grid(row=5, column=0, columnspan=2, padx=10, pady=10)
    dont = Label(up, text="Already have an account?", font=("HELVETICA", 12), bg="#f8f8f8")
    dont.grid(row=6, column=0, columnspan=2, padx=(0, 100))
    login_btn = Button(up, text='Login', bd=3, relief=RAISED, width=10, font=("HELVETICA", 12), command=log_back)
    login_btn.grid(row=6, column=1, padx=(90, 0), pady=10)

    
  

def home():
    def logout():
        home_w.destroy()
        log.deiconify()
        username.delete(0, END)
        password.delete(0, END)
    def grade():
        grade = Toplevel()
        grade.title('Grade Book')
        grade.resizable(0, 0)
        grade.configure(bg="#f8f8f8")
        heading = Label(grade, text=username.get().upper() + ' SCORES', font=("HELVETICA", 20), bg="#f8f8f8")
        heading.pack(fill=BOTH, expand=YES, pady=10)
        grade_score = ttk.Treeview(grade, columns=('Quiz Name', 'Score', 'Percentage'), show="headings")
        grade_score.heading('#1', text='Quiz Name', anchor=CENTER)
        grade_score.heading('#2', text='Score', anchor=CENTER)
        grade_score.heading('#3', text='Percentage', anchor=CENTER)
        with open('user_score.csv') as f:
            reader = csv.DictReader(f, delimiter=',')
            for row in reader:
                name = row['Student_Name']
                if username.get() == name:
                    head = reader.fieldnames
                    for i in range(1, len(head)):
                        q_score = row[head[i]]
                        if q_score != '':
                            percent = (int(q_score) / 10) * 100
                            grade_score.insert("", END, values=(head[i], q_score, str(percent) + '%'))
                        else:
                            grade_score.insert("", END, values=(head[i], 0, '0%'))
        scroll = Scrollbar(grade, command=grade_score.yview, orient='vertical')
        grade_score.configure(yscrollcommand=scroll.set)
        scroll.pack(side=RIGHT, fill=Y)
        grade_score.pack(fill=BOTH, expand=YES, pady=(0, 20), padx=10)
    def main():
        def back():
            root.destroy()
            home_w.deiconify()
        def show_quiz(name):
            def back2():
                root2.destroy()
                home_w.deiconify()
            def start_quiz(f, quiz_num, head):
                root2.withdraw()
                global counter
                counter = 0
                data = []
                def check(b):
                    global counter, score, all_files
                    answer = b.cget('text')
                    correct_ans = data[counter][6][1]
                    correct = data[counter][int(correct_ans) + 1]
                    if str(correct) == str(answer):
                        score += 1
                        cor_wron.append(f'Your answer is correct.')
                    else:
                        cor_wron.append(f'Your answer is wrong. The Correct answer is {str(correct)}')
                    try:
                        if counter < 10:
                            counter += 1
                            num['text'] = 'Q' + data[counter][0]
                            quest['text'] = data[counter][1]
                            option1.configure(text=data[counter][2])
                            option2.configure(text=data[counter][3])
                            option3.configure(text=data[counter][4])
                            option4.configure(text=data[counter][5])
                    except:
                        try:
                            nn = str(name[0]).upper() + str(name[1]).upper() + str(name[2]).upper() + str(name[3]).upper()
                            quiz_score[nn] = score
                        except:
                            nn = str(name[0]).upper() + str(name[1]).upper() + str(name[2]).upper()
                            quiz_score[nn] = score
                        root3.destroy()
                        root2.deiconify()
                        score = 0
                        result = Toplevel()
                        result.title('Quiz')
                        result.resizable(0, 0)
                        result.configure(bg="#f8f8f8")
                        f_score = Label(result, text=f'Your Quiz Score is: {quiz_score[nn]} out of 5',
                                        font=("HELVETICA", 14, 'bold', 'underline'), bg="#f8f8f8")
                        f_score.grid(row=1, column=1, padx=40, pady=20)
                        for i in range(len(cor_wron)):
                            ri_wro_label = Label(result, text=f'{i + 1}. {cor_wron[i]}',
                                                 font=("HELVETICA", 14), bg="#f8f8f8")
                            ri_wro_label.grid(row=2 + i, column=1, padx=40, pady=20)
                        f = open('user_score.txt', 'a', encoding='UTF8', newline='')
                        f.write(f'\n{username.get()}, {name.title()}, {quiz_score[nn]}, 5')
                        f.close()
                        cor_wron.clear()
                        result.mainloop()
                        lst = list(quiz_score.items())[-1]
                        quiz_score.clear()
                conn = sqlite3.connect('question_bank.db')
                cursor = conn.cursor()
                cursor.execute(f"Select * from [{name.lower()}]")
                conn.commit()
                a = cursor.fetchall()
                c = []
                for i in a:
                    c.append(list(i))
                data = c
                for i in range(len(data)):
                    data[i][0] = str(data[i][0])
                random.shuffle(data)
                if len(data) >= 5:
                    data = random.sample(data, 5)
                root3 = Toplevel()
                root3.title('Quiz')
                root3.resizable(0, 0)
                h = root3.winfo_screenheight()
                w = root3.winfo_screenwidth()
                root3.geometry('{}x{}'.format(int(w) - 100, int(h / 2) - 50))
                answer = StringVar()
                num = Label(root3, text='Q' + data[counter][0], font=("HELVETICA", 14))
                num.grid(row=0, column=0, padx=(0, 10), pady=10)
                quest = Label(root3, text=data[counter][1], font=("HELVETICA", 14))
                quest.grid(row=0, column=1, columnspan=2, padx=10, pady=10)
                option1 = Button(root3, text=data[counter][2], bg='red', width=30, bd=3, relief=RAISED,
                                 font=("HELVETICA", 12), wraplength=250, justify=CENTER)
                option1.configure(command=lambda b=option1: check(b))
                option1.grid(row=1, column=1, padx=10, pady=(10, 5))
                option2 = Button(root3, text=data[counter][3], bg='lightblue', width=30, bd=3, relief=RAISED,
                                 font=("HELVETICA", 12), wraplength=250, justify=CENTER)
                option2.configure(command=lambda b=option2: check(b))
                option2.grid(row=1, column=2, padx=10, pady=(10, 5))
                option3 = Button(root3, text=data[counter][4], bg='orange', width=30, bd=3, relief=RAISED,
                                 font=("HELVETICA", 12), wraplength=250, justify=CENTER)
                option3.configure(command=lambda b=option3: check(b))
                option3.grid(row=2, column=1, padx=10, pady=(10, 5))
                option4 = Button(root3, text=data[counter][5], bg='pink', width=30, bd=3, relief=RAISED,
                                 font=("HELVETICA", 12), wraplength=250, justify=CENTER)
                option4.configure(command=lambda b=option4: check(b))
                option4.grid(row=2, column=2, padx=10, pady=(10, 5))
                root3.columnconfigure(3, weight=1)
                root3.rowconfigure(3, weight=1)
                root3.grid_columnconfigure(3, weight=1)
                root3.grid_rowconfigure(3, weight=1)
                root3.mainloop()
            root.withdraw()
            root2 = Toplevel()
            root2.title('Quiz Select')
            root2.resizable(0, 0)
            lbl1 = Label(root2, text=name.upper(), font=("HELVETICA", 14))
            lbl1.grid(row=0, column=0, columnspan=2, padx=30, pady=20)
            select = Label(root2, text='Select a Quiz', font=("HELVETICA", 14))
            select.grid(row=1, column=0, padx=10, pady=5)
            global all_files
            all_files = [f for f in os.listdir('.') if '.db' in f]
            head = ['Student_Name']
            try:
                head.append(name[0].upper() + name[1].upper() + name[2].upper() + name[3].upper())
            except:
                head.append(name[0].upper() + name[1].upper() + name[2].upper())
            num_files1 = [f for f in os.listdir('.') if name.lower() in f]
            conn = sqlite3.connect('question_bank.db')
            cursor = conn.cursor()
            cursor.execute(f"Select * from [{name.lower()}]")
            conn.commit()
            a = cursor.fetchall()
            c = []
            for i in a:
                c.append(list(i))
            data = c
            for i in range(len(data)):
                data[i][0] = str(data[i][0])
            r = 1
            btn = Button(root2, text='Quiz', width=10, bd=3, relief=RAISED, font=("HELVETICA", 14),
                         command=lambda head='a', quiz_num=1, f=1: start_quiz(f, quiz_num, head))
            btn.grid(row=r, column=1, pady=5, padx=10)
            r += 1
            menu = Button(root2, text='Menu', width=5, font=("HELVETICA", 14), command=back2)
            menu.grid(row=r, column=0, padx=(0, 90), pady=(15, 0))
            root2.mainloop()
        home_w.withdraw()
        root = Toplevel()
        root.title('Quiz Subject')
        root.resizable(0, 0)
        lbl = Label(root, text='Pick a Subject', font=("HELVETICA", 15))
        lbl.grid(row=0, column=0, columnspan=2, padx=40, pady=20)
        con3 = sqlite3.connect("question_bank.db")
        cursor3 = con3.cursor()
        cursor3.execute("SELECT name FROM sqlite_master WHERE type='table';")
        con3.commit()
        aaa = cursor3.fetchall()
        ccc = []
        for i in aaa:
            ccc.append(list(i))
        data2 = ccc
        for i in range(len(data2)):
            data2[i][0] = str(data2[i][0])
        new_row = 1
        for i in range(len(data2)):
            btn = Button(root, text=f'{data2[i][0].title()}', width=35, font=("HELVETICA", 14),
                         command=lambda name=f'{data2[i][0].title()}': show_quiz(name))
            btn.grid(row=new_row, column=0, columnspan=2, padx=40, pady=5)
            new_row += 1
        menu = Button(root, text='Menu', width=8, font=("HELVETICA", 14), command=back)
        menu.grid(row=new_row, column=0, padx=(0, 150), pady=(15, 0))
        root.mainloop()
    log.withdraw()
    home_w = Toplevel()
    home_w.resizable(0, 0)
    home_w.title('Home Page')
    home_w.configure(bg="#f8f8f8")
    logout_btn = Button(home_w, text='Logout', font=("ARIAL", 15), bd=3, relief=RAISED, command=logout)
    logout_btn.grid(row=0, column=1, padx=(30, 0), pady=(0, 10))
    head = Label(home_w, text='WELCOME TO\nPROJECT QUIZ', font=("HELVETICA", 20, 'bold'), bg="#f8f8f8")
    head.grid(row=1, column=0, columnspan=2, padx=20, pady=20)
    play_btn = Button(home_w, text='Play', font=("ARIAL", 15), bd=3, relief=RAISED, command=main)
    play_btn.grid(row=2, column=0, padx=(10, 30), pady=(10, 30))
    home_w.mainloop()

from tkinter import Tk, Label, Entry, Button, Frame
from PIL import Image, ImageTk, ImageDraw  

# Create the main window
log = Tk()
log.title('Login Page')
log.attributes('-fullscreen', True)


# Set window size and align it toward the right
w, h = 650, 750  
ws, hs = log.winfo_screenwidth(), log.winfo_screenheight()
x, y = ws - w - 50, (hs - h) // 2  
log.geometry(f"{w}x{h}+{x}+{y}")
log.resizable(True, True)
log.configure(bg="#E5D9F2")  

# Create the form frame
frame = Frame(log, bg="white", padx=50, pady=50, relief="raised", bd=5)
frame.pack(pady=50)

# **Load and Convert Logo Image into a Circle**
def get_circular_image(image_path, size=(160, 160)):  
    try:
        img = Image.open(image_path).convert("RGBA")  
        img = img.resize(size, Image.Resampling.LANCZOS)  

        # Create a circular mask
        mask = Image.new("L", size, 0)
        draw = ImageDraw.Draw(mask)
        draw.ellipse((0, 0) + size, fill=255)

        # Apply the mask to the image
        circular_img = Image.new("RGBA", size, (0, 0, 0, 0))
        circular_img.paste(img, (0, 0), mask)

        return ImageTk.PhotoImage(circular_img)  

    except Exception as e:
        print("Error loading image:", e)
        return None

# Load the circular logo
logo_photo = get_circular_image("education.png")  
if logo_photo:
    logo_label = Label(frame, image=logo_photo, bg="white")
    logo_label.grid(row=0, column=0, columnspan=2, pady=10)

# **Header Label**
heading = Label(frame, text='PROJECT QUIZ', font=("Georgia", 24, 'bold'), bg="white", fg="#003366")
heading.grid(row=1, column=0, columnspan=2, pady=15)

# **Username Input**
username_lbl = Label(frame, text='Username:', font=("Arial", 16, 'bold'), bg="white", fg="#333333")
username_lbl.grid(row=2, column=0, sticky="w", pady=10)
username = Entry(frame, width=30, font=("Arial", 16), bd=2, relief="solid")
username.grid(row=2, column=1, pady=10)

# **Password Input**
password_lbl = Label(frame, text='Password:', font=("Arial", 16, 'bold'), bg="white", fg="#333333")
password_lbl.grid(row=3, column=0, sticky="w", pady=10)
password = Entry(frame, width=30, show="*", font=("Arial", 16), bd=2, relief="solid")
password.grid(row=3, column=1, pady=10)

# **Button Styling**
def on_enter(event, btn, color):
    btn.config(bg=color)

def on_leave(event, btn, original_color):
    btn.config(bg=original_color)
def close_app():
    log.destroy()

# Create a Close Button
close_btn = Button(log, text="X", font=("Arial", 16, "bold"), bg="red", fg="white",
                   bd=2, relief="raised", command=close_app,
                   activebackground="darkred", activeforeground="white", highlightthickness=0)
close_btn.place(x=log.winfo_screenwidth() - 50, y=10, width=40, height=40)  # Position at the top right

# Escape Key to Close
log.bind("<Escape>", lambda event: close_app())

# **Login Button**
login_btn = Button(frame, text='Login', font=('Arial', 16, 'bold'), width=24, bg="#008CBA", fg="white",
                   bd=4, relief="raised", command=login,
                   activebackground="#005f7f", activeforeground="white", highlightthickness=0)
login_btn.grid(row=4, column=0, columnspan=2, pady=20)
login_btn.bind("<Enter>", lambda e: on_enter(e, login_btn, "#005f7f"))
login_btn.bind("<Leave>", lambda e: on_leave(e, login_btn, "#008CBA"))

# **Signup & Admin Section**
btn_frame = Frame(frame, bg="white")
btn_frame.grid(row=5, column=0, columnspan=2, pady=10)

# "Don't have an account?" Label
dont_have_label = Label(btn_frame, text="Don't have an account?", font=("Arial", 14), bg="white", fg="black")
dont_have_label.grid(row=0, column=0, padx=(0, 5), sticky="e")

# **Sign Up Button**
signup_btn = Button(btn_frame, text='Sign Up', font=("Arial", 14, 'bold'), width=12, bg="#4CAF50", fg="white",
                    bd=2, relief="raised", command=signup,
                    activebackground="#357a38", activeforeground="white", highlightthickness=0)
signup_btn.grid(row=0, column=1, padx=(0, 10))

signup_btn.bind("<Enter>", lambda e: on_enter(e, signup_btn, "#357a38"))
signup_btn.bind("<Leave>", lambda e: on_leave(e, signup_btn, "#4CAF50"))

# "Are you an Admin?" Label
admin_label = Label(btn_frame, text="Are you an Admin?", font=("Arial", 14), bg="white", fg="black")
admin_label.grid(row=1, column=0, padx=(0, 5), sticky="e", pady=10)

# **Admin Login Button**
admin_btn = Button(btn_frame, text='Admin Login', font=("Arial", 14, 'bold'), width=12, bg="#FF5733", fg="white",
                   bd=2, relief="raised", command=admin_log,
                   activebackground="#a82e1e", activeforeground="white", highlightthickness=0)
admin_btn.grid(row=1, column=1, padx=(0, 10), pady=10)

admin_btn.bind("<Enter>", lambda e: on_enter(e, admin_btn, "#a82e1e"))
admin_btn.bind("<Leave>", lambda e: on_leave(e, admin_btn, "#FF5733"))

# Run the application
log.mainloop()
